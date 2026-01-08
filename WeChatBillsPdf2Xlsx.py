import datetime
import tkinter as tk
from tkinter import messagebox
import tkinter.ttk as ttk
from tkinter import filedialog, LEFT, StringVar
import windnd
import os
import pathlib
import pdfplumber
import openpyxl
from threading import Thread


def convert_pdf(files, target):
    try:
        total_rows = []
        for fi, file in enumerate(files):
            with pdfplumber.open(file) as pdf:
                for pi, page in enumerate(pdf.pages):
                    tables = page.extract_tables()
                    for ti, table in enumerate(tables):
                        if fi != 0 and pi == 0 and ti == 0:
                            table = table[3:]  # 去除重复表头
                        for row in table:
                            if len(row) > 1:
                                row[1] = row[1].replace("\n", " ") if row[1] and "\n" in row[1] else row[1]
                                row = [col.replace("\n", "") if col else "" for col in row]
                                total_rows.append(row)
        if len(total_rows) < 4:
            messagebox.showwarning("警告", "文件中未找到表格")
            return

        header = total_rows[2]
        total_rows = total_rows[3:]
        total_rows.sort(key=lambda ele: ele[1])  # 按日期排序
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "账单"
        for ci, col in enumerate(header):  # 写表头
            sheet.cell(row=1, column=ci + 1, value=col)
        for ri, row in enumerate(total_rows):
            for ci, col in enumerate(row):
                if ci == 1:  # 日期转格式
                    try:
                        sheet.cell(row=ri + 2, column=ci + 1, value=datetime.datetime.fromisoformat(col))
                        sheet.cell(row=ri + 2, column=ci + 1).number_format = 'yyyy-mm-dd hh:mm:ss'
                    except:
                        pass
                elif ci == 5:  # 金额转小数
                    try:
                        sheet.cell(row=ri + 2, column=ci + 1, value=float(col))
                        sheet.cell(row=ri + 2, column=ci + 1).number_format = '0.00'  # 保留两位小数
                    except:
                        pass
                else:
                    sheet.cell(row=ri + 2, column=ci + 1, value=col)
        fn = os.path.join(target, f"convert_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx")
        wb.save(fn)
        messagebox.showinfo("提示", f"转换成功\n共{len(total_rows)}条记录\n已保存至{fn}")

    except Exception as e:
        messagebox.showerror("错误", f"转换失败:\n{e}")


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("微信账单PDF转Excel")
        self.geometry('720x100')
        self.resizable(False, False)
        self.iconbitmap(os.path.join(pathlib.Path(__file__).parent.resolve(), "icon.ico"))
        file_frame = ttk.Frame(self)
        file_frame.pack(padx=5, pady=5)
        ttk.Label(file_frame, text="目录：").pack(side=LEFT, padx=5, pady=5)
        self.target_folder = StringVar()
        ttk.Entry(file_frame, width=60, textvariable=self.target_folder).pack(side=LEFT, padx=5, pady=5)
        ttk.Button(file_frame, text="选择", command=self.select_folder).pack(side=LEFT, padx=5, pady=5)
        action_frame = ttk.Frame(self)
        action_frame.pack(padx=5, pady=5)
        ttk.Button(action_frame, text="转换", width=20, command=self.start_convert).pack(side=LEFT, padx=5, pady=5)
        self.conv_process = None

    def dragged_files(self, files):
        if len(files):
            target = files[0].decode("gbk")
            self.target_folder.set(target if os.path.isdir(target) else os.path.dirname(target))

    def select_folder(self):
        tdir = tk.filedialog.askdirectory()
        self.target_folder.set(tdir.replace("/", "\\")) if tdir else None

    def start_convert(self):
        if self.conv_process and self.conv_process.is_alive():
            messagebox.showwarning("警告", "正在转换，请稍等...")
            return
        target = self.target_folder.get()
        files = [os.path.join(target, f.name) for f in (pathlib.Path(target).glob('*.pdf'))]
        if not files:
            messagebox.showwarning("警告", "未找到pdf文件")
            return
        self.conv_process = Thread(target=convert_pdf, args=(files, target))
        self.conv_process.start()

    def on_close(self):
        if self.conv_process and self.conv_process.is_alive():
            messagebox.showwarning("警告", "正在转换，请稍等...")
            return
        self.destroy()


if __name__ == '__main__':
    app = App()
    app.protocol("WM_DELETE_WINDOW", app.on_close)
    windnd.hook_dropfiles(app, func=app.dragged_files)
    app.mainloop()
